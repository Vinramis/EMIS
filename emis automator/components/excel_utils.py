
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

def get_cell_value(sheet, cell_ref: str):
    """
    Gets cell value from Worksheet by reference (e.g., 'A1').
    """
    try:
        return sheet[cell_ref].value
    except Exception:
        return None

def generate_sequence(sheet, start_cell: str, mode: str, max_empty_cells_in_a_row: int = 5) -> list[str]:
    """
    Generates a sequence of non-empty cell references from Worksheet.
    """
    sequence = []
    empty_cells_in_a_row = 0

    try:
        # Parse start_cell (e.g., 'A1') into column letter and row number
        xy = coordinate_from_string(start_cell) 
        current_col_letter = xy[0]
        current_row = xy[1]
        
        # Convert column letter to index (1-based) for easier manipulation
        current_col_idx = column_index_from_string(current_col_letter)
        
    except ValueError:
        return []

    while empty_cells_in_a_row < max_empty_cells_in_a_row:
        current_col_letter = get_column_letter(current_col_idx)
        cell_ref = f"{current_col_letter}{current_row}"
        
        cell_value = get_cell_value(sheet, cell_ref)
        # Check if empty or too short (similar to original logic)
        is_empty = cell_value is None or len(str(cell_value).strip()) < 2
        
        if is_empty:
            empty_cells_in_a_row += 1
        else:
            empty_cells_in_a_row = 0
            sequence.append(cell_ref)

        if mode == "col":
            current_row += 1
        elif mode == "row":
            current_col_idx += 1
        else:
            break

    return sequence

def read_topics_from_excel(file_path: str, start_cell: str, mode: str, starting_topic_number: int = 1, ending_topic_number: int = -1) -> list[str]:
    """
    Helper to read topics from the excel file.
    """
    try:
        # Load workbook and select active sheet
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        cell_sequence = generate_sequence(sheet, start_cell, mode)
        
        values = []
        for cell_ref in cell_sequence:
            val = get_cell_value(sheet, cell_ref)
            if val is not None:
                values.append(str(val).strip())
                
        if ending_topic_number == -1:
            return values
        else:
            return values[starting_topic_number - 1:ending_topic_number - 1]
    except Exception as e:
        print(f"[ОШИБКА] Чтение файла Excel: {e}")
        return []
